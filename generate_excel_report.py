"""
generate_excel_report.py  —  Data Quality Governance Project
Generates a formatted Excel governance report from CSV exports.

Run after export_to_csv.py to ensure CSVs are up to date.

Author: Feranmi Okunola
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ── File paths ───────────────────────────────────────────────────────────────
BASE   = r"C:\Users\flofa\OneDrive\Data Quality Governance Project"
OUTPUT = os.path.join(BASE, "Data_Governance_Report.xlsx")

cde    = pd.read_csv(os.path.join(BASE, "cde_inventory.csv"))
ctrl   = pd.read_csv(os.path.join(BASE, "control_register.csv"))
scores = pd.read_csv(os.path.join(BASE, "score_log.csv"))

# ── Sheet 1: CDE Summary ─────────────────────────────────────────────────────
cde_summary = cde.merge(scores, on="cde_id", how="left")[[
    "cde_name", "division", "system_name", "criticality",
    "overall_score", "risk_classification"
]].sort_values("risk_classification")

# ── Sheet 2: Active Breach Register ──────────────────────────────────────────
breaches = ctrl[ctrl["control_status"].isin(["Breached", "Missing"])][[
    "cde_name", "dimension", "control_status", "division",
    "system_name", "breach_date", "sla_deadline"
]].sort_values(["division", "cde_name"])

# ── Sheet 3: Control Status by Division ──────────────────────────────────────
division_pivot = ctrl.pivot_table(
    index="division",
    columns="control_status",
    values="control_id",
    aggfunc="count",
    fill_value=0
).reset_index()
division_pivot.columns.name = None

# ── Write to Excel ────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    cde_summary.to_excel(writer,    sheet_name="CDE Summary",                index=False)
    breaches.to_excel(writer,       sheet_name="Active Breach Register",     index=False)
    division_pivot.to_excel(writer, sheet_name="Control Status by Division", index=False)

# ── Apply formatting ──────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT)
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

wb.save(OUTPUT)
print(f"Report saved to: {OUTPUT}")
